# -*- coding: utf-8 -*-
"""
gerar_dados_boasafra.py
=======================
Gera os dados brutos da Comercial Boa Safra como sairiam de CADA sistema,
no formato/dialeto nativo de cada um — reproduzindo de propósito a
heterogeneidade e as chaves incompatíveis descritas no diagnóstico.

Uma "verdade" oculta (clientes e produtos reais) alimenta os quatro sistemas,
mas cada sistema representa a mesma realidade de um jeito diferente:

  ERP (Protheus)   -> CSV Latin-1 ';' + TXT posicional | chave COD_CLI numérico
  Força de vendas  -> Excel (.xlsx)                     | chave CLIENTE_ID "UF-nnnn" + CNPJ
  E-commerce B2B   -> CSV UTF-8 moderno                 | chave e-mail do comprador + EAN
  Planilhas (manual)-> Excel bagunçado                  | chave nome digitado à mão

Saída: dados_sistemas/
"""
import csv
import unicodedata
import numpy as np
import pandas as pd
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

RNG = np.random.default_rng(7)
OUT = "dados_sistemas"
INI = date(2025, 1, 1)
FIM = date(2026, 6, 30)          # 18 meses, terminando recente
N_CLI = 500
N_PROD = 600

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def strip_acc(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def cnpj_check(base12):
    def dv(nums, pesos):
        s = sum(int(n)*p for n, p in zip(nums, pesos))
        r = s % 11
        return "0" if r < 2 else str(11 - r)
    p1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    p2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]
    d1 = dv(base12, p1)
    d2 = dv(base12 + d1, p2)
    return base12 + d1 + d2

def gen_cnpj():
    base = f"{RNG.integers(10000000,99999999)}0001"   # matriz 0001
    return cnpj_check(base)                            # 14 dígitos

def fmt_cnpj(d):
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"

def ean13(base12):
    s = sum((int(n) * (3 if i % 2 else 1)) for i, n in enumerate(base12))
    return base12 + str((10 - s % 10) % 10)

def money_br(v):      # 1234.5 -> "1234,56"
    return f"{v:.2f}".replace(".", ",")

def baguncar_nome(razao, fantasia):
    """Simula um nome digitado à mão numa planilha: abreviações, minúsculas,
    perda de acento, sufixo solto — a origem do inferno de VLOOKUP."""
    base = fantasia if (fantasia and RNG.random() < 0.6) else razao
    s = base
    if RNG.random() < 0.5: s = s.title()
    if RNG.random() < 0.4: s = s.lower()
    s = s.replace("Comércio", "Com.").replace("Comercio", "Com.")
    s = s.replace("Distribuidora", "Distrib.").replace("Mercado", "Merc.")
    s = s.replace(" LTDA", "").replace(" Ltda", "").replace(" - ME", "").replace(" EIRELI", "")
    if RNG.random() < 0.35: s = strip_acc(s)
    if RNG.random() < 0.15 and len(s) > 6:                 # typo ocasional
        i = RNG.integers(1, len(s)-1); s = s[:i] + s[i+1:]
    return s.strip()

# ---------------------------------------------------------------------------
# VERDADE OCULTA — produtos
# ---------------------------------------------------------------------------
CATS = {
    "Mercearia Seca":  (3, 40, 1.30, 1.65),
    "Bebidas":         (5, 30, 1.25, 1.55),
    "Limpeza":         (4, 35, 1.35, 1.90),
    "Higiene Pessoal": (6, 45, 1.40, 1.95),
    "Descartáveis":    (3, 22, 1.30, 1.70),
    "Matinais":        (4, 50, 1.30, 1.75),
}
NOMES_PROD = {
    "Mercearia Seca": ["Arroz T1 5kg","Feijão Carioca 1kg","Açúcar Refinado 1kg","Óleo de Soja 900ml",
        "Macarrão Espaguete 500g","Farinha de Trigo 1kg","Sal Refinado 1kg","Café Torrado 500g","Molho de Tomate 340g"],
    "Bebidas": ["Refrigerante Cola 2L","Suco Néctar 1L","Água Mineral 1,5L","Cerveja Lata 350ml",
        "Energético 250ml","Refresco em Pó 30g","Isotônico 500ml"],
    "Limpeza": ["Detergente Neutro 500ml","Água Sanitária 2L","Sabão em Pó 1kg","Amaciante 2L",
        "Desinfetante 1L","Esponja Multiuso","Limpador Multiuso 500ml","Sabão em Barra 200g"],
    "Higiene Pessoal": ["Sabonete 90g","Shampoo 350ml","Creme Dental 90g","Papel Higiênico 4un",
        "Desodorante Aerossol","Absorvente 8un","Fralda Infantil M"],
    "Descartáveis": ["Copo Descartável 200ml","Prato Descartável 15cm","Guardanapo 50un",
        "Saco de Lixo 50L","Papel Toalha 2un","Talher Descartável"],
    "Matinais": ["Leite Integral 1L","Achocolatado 400g","Biscoito Recheado 130g","Margarina 500g",
        "Pão de Forma 500g","Cereal Matinal 300g","Iogurte 170g"],
}
produtos = []
pcod = 0
for cat, (cmin, cmax, mk_lo, mk_hi) in CATS.items():
    marcas = ["Bom Preço","Aurora","Vitória","Primor","Delta","Real","Campo Verde"]
    for _ in range(N_PROD // len(CATS)):
        pcod += 1
        base_nome = RNG.choice(NOMES_PROD[cat])
        marca = RNG.choice(marcas)
        custo = round(float(RNG.uniform(cmin, cmax)), 2)
        preco = round(custo * float(RNG.uniform(mk_lo, mk_hi)), 2)
        produtos.append({
            "id": pcod,
            "cod_prod": f"PA{pcod:05d}",                       # código ERP Protheus
            "ean": ean13(f"789{RNG.integers(100000000,999999999)}"),
            "descricao": f"{base_nome} {marca}",
            "categoria": cat,
            "unidade": RNG.choice(["UN","CX","FD","PC"]),
            "custo": custo,
            "preco": preco,
        })
prod_by_id = {p["id"]: p for p in produtos}

# ---------------------------------------------------------------------------
# VERDADE OCULTA — clientes
# ---------------------------------------------------------------------------
REGioes = {  # UF -> (municípios, peso)
    "MG": (["Contagem","Betim","Belo Horizonte","Uberlândia","Juiz de Fora","Sete Lagoas"], 0.42),
    "GO": (["Goiânia","Anápolis","Aparecida de Goiânia","Rio Verde"], 0.24),
    "SP": (["Ribeirão Preto","Franca","São Paulo"], 0.14),
    "DF": (["Brasília","Taguatinga"], 0.10),
    "BA": (["Barreiras","Luís Eduardo Magalhães"], 0.06),
    "ES": (["Vila Velha","Serra"], 0.04),
}
SEGMENTOS = ["Mercado de Bairro","Padaria","Food Service","Mini Rede","Conveniência","Bar/Restaurante"]
TIPOS_RS = ["Comércio de Alimentos","Distribuidora","Mercado","Supermercado","Comercial","Empório"]
ufs = list(REGioes)
uf_p = np.array([REGioes[u][1] for u in ufs]); uf_p = uf_p / uf_p.sum()

clientes = []
app_seq = {u: 1000 for u in ufs}
for i in range(1, N_CLI + 1):
    uf = RNG.choice(ufs, p=uf_p)
    mun = RNG.choice(REGioes[uf][0])
    seg = RNG.choice(SEGMENTOS)
    perfil = RNG.choice(["top","medio","dreno"], p=[0.22, 0.55, 0.23])
    porte = float(RNG.gamma(2.0, 1.0))
    nome_base = RNG.choice(["Silva","Souza","Oliveira","Pereira","Costa","Almeida","Ferreira",
                            "Nogueira","Bittencourt","do Vale","Andrade","Rezende"])
    razao = f"{RNG.choice(TIPOS_RS)} {nome_base} {RNG.choice(['LTDA','LTDA','EIRELI','- ME'])}"
    fantasia = RNG.choice(["Merc. Central","Empório Bom Dia","Padaria Estrela","Rei do Atacado",
                           "Mercado do Zé","Casa da Cesta","Super Econômico","Mercadinho Popular",
                           "Distribuidora Ouro","Ponto Certo"]) + f" {i:03d}"
    # canais que o cliente usa (representante / e-commerce / balcão)
    canais = set(["erp"])                      # todos faturam no ERP
    if RNG.random() < 0.62: canais.add("rep")
    if RNG.random() < 0.40: canais.add("ecom")
    if not (canais - {"erp"}): canais.add("rep")
    cnpj = gen_cnpj()
    clientes.append({
        "id": i,
        "cod_cli": f"{i:06d}",                                   # chave ERP (numérica)
        "cliente_id_app": f"{uf}-{app_seq[uf]:04d}" if "rep" in canais else "",
        "email": (f"compras@{strip_acc(nome_base).lower()}{i}.com.br" if "ecom" in canais else ""),
        "cnpj": cnpj,
        "razao": razao.upper(),
        "fantasia": fantasia,
        "uf": uf, "municipio": mun, "segmento": seg,
        "perfil": perfil, "porte": porte,
        "rep": None, "canais": canais,
    })
    if "rep" in canais: app_seq[uf] += 1

# representantes (força de vendas)
REPS = [f"{n} {s}" for n, s in zip(
    ["Carlos","Marcos","Fernanda","Rafael","Juliana","André","Patrícia","Diego","Luana","Rodrigo"],
    ["Menezes","Tavares","Lima","Braga","Cardoso","Pinto","Ramos","Freitas","Barros","Moura"])]
for c in clientes:
    if "rep" in c["canais"]:
        c["rep"] = REPS[RNG.integers(0, len(REPS))]

cli_com_ecom = [c for c in clientes if "ecom" in c["canais"]]

# ---------------------------------------------------------------------------
# GERA PEDIDOS (verdade) e projeta em cada sistema
# ---------------------------------------------------------------------------
dias = [INI + timedelta(days=d) for d in range((FIM - INI).days + 1)]
pesos_cli = np.array([c["porte"] for c in clientes]); pesos_cli /= pesos_cli.sum()

erp_linhas, app_pedidos, ecom_linhas = [], [], []
nota_seq = {"01": 100000, "02": 500000}      # filiais MG / GO
ped_app_seq = 70000
ecom_seq = 900000

def filial_do(uf):
    return "02" if uf in ("GO","DF") else "01"

for d in dias:
    t = (d - INI).days
    tend = 1.0 + 0.0004 * t
    saz_ano = 1.0 + 0.25 * np.sin(2*np.pi*(d.timetuple().tm_yday)/365 - 0.4)
    sw = {0:1.05,1:1.0,2:1.0,3:1.05,4:1.2,5:0.65,6:0.3}[d.weekday()]
    n = RNG.poisson(48 * tend * saz_ano * sw)
    idxs = RNG.choice(len(clientes), size=n, p=pesos_cli)
    for ci in idxs:
        c = clientes[ci]
        opc = [k for k in ("rep","ecom") if k in c["canais"]]
        if not opc:
            canal = "balcao"
        else:
            canal = RNG.choice(opc + (["balcao"] if RNG.random() < 0.25 else []))
        p = prod_by_id[int(RNG.integers(1, N_PROD+1))]
        base_q = RNG.gamma(2.0, 2.5)
        if c["perfil"] == "top":   qtd = int(max(1, base_q*(1.3+0.4*c["porte"])))
        elif c["perfil"] == "medio": qtd = int(max(1, base_q*(0.8+0.3*c["porte"])))
        else:                       qtd = int(max(1, base_q*0.4))
        if c["perfil"] == "dreno":  desc = RNG.choice([0.05,0.1,0.15,0.2], p=[.2,.35,.3,.15])
        elif c["perfil"] == "top":  desc = RNG.choice([0,0,0.05], p=[.7,.2,.1])
        else:                       desc = RNG.choice([0,0.05,0.1], p=[.6,.25,.15])
        preco = round(p["preco"]*(1-desc), 2)
        fil = filial_do(c["uf"])
        nota_seq[fil] += 1
        nota = nota_seq[fil]
        # ---- ERP faturamento (todas as vendas) ----
        erp_linhas.append({
            "FILIAL": fil, "NUMNOTA": f"{nota:06d}", "SERIE": "1",
            "EMISSAO": d.strftime("%d/%m/%Y"),
            "CODCLI": c["cod_cli"], "CODPROD": p["cod_prod"],
            "QUANT": qtd, "VLRUNIT": money_br(preco),
            "VLRTOTAL": money_br(round(qtd*preco, 2)),
            "CFOP": RNG.choice(["5102","6102"]), "COND_PGTO": RNG.choice(["28 DD","30 DD","À VISTA","28/35/42"]),
        })
        # ---- Força de vendas (só canal representante) ----
        if canal == "rep":
            ped_app_seq += 1
            comis = round(float(RNG.uniform(1.5, 3.5)), 2)
            app_pedidos.append({
                "Pedido": f"PV{ped_app_seq}", "Data": d,
                "CLIENTE_ID": c["cliente_id_app"], "CNPJ": c["cnpj"],
                "Cliente": c["fantasia"], "Representante": c["rep"],
                "Produto": p["descricao"], "Qtde": qtd,
                "Preço Unit.": preco, "Total": round(qtd*preco, 2),
                "Comissão %": comis, "UF": c["uf"],
            })
        # ---- E-commerce (só canal online) ----
        if canal == "ecom":
            ecom_seq += 1
            nome_digitado = c["fantasia"] if RNG.random() < 0.5 else baguncar_nome(c["razao"], c["fantasia"])
            if RNG.random() < 0.03: nome_digitado = ""       # comprador não preencheu
            ecom_linhas.append({
                "order_id": f"WEB-{ecom_seq}", "order_date": d.strftime("%Y-%m-%d"),
                "customer_email": c["email"], "customer_name": nome_digitado,
                "product_ean": p["ean"], "product_name": p["descricao"],
                "qty": qtd, "unit_price": round(preco, 2),
                "payment_method": RNG.choice(["boleto","pix","cartao_credito"]),
                "channel": "b2b_web",
            })

# ---- injeta imperfeições reais no ERP faturamento ----
erp_df = pd.DataFrame(erp_linhas)
# ~2% de vendas com COD_CLI que não existe no cadastro (cadastro incompleto)
mask = RNG.random(len(erp_df)) < 0.02
erp_df.loc[mask, "CODCLI"] = [f"9{RNG.integers(10000,99999)}" for _ in range(mask.sum())]

# ---------------------------------------------------------------------------
# ESCRITA — cada sistema no seu formato nativo
# ---------------------------------------------------------------------------
import os
os.makedirs(OUT, exist_ok=True)

# 1) ERP_FATURAMENTO.csv  — Latin-1, ';', decimal ',', DD/MM/YYYY, MAIÚSCULAS
with open(f"{OUT}/ERP_FATURAMENTO.csv", "w", encoding="latin-1", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(erp_df.columns), delimiter=";")
    w.writeheader()
    for _, r in erp_df.iterrows():
        w.writerow(r.to_dict())

# 2) ERP_CLIENTES.csv — Latin-1, ';', CNPJ mascarado
with open(f"{OUT}/ERP_CLIENTES.csv", "w", encoding="latin-1", newline="") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(["COD_CLI","RAZAO_SOCIAL","CNPJ","UF","MUNICIPIO","SEGMENTO"])
    for c in clientes:
        w.writerow([c["cod_cli"], c["razao"], fmt_cnpj(c["cnpj"]), c["uf"], c["municipio"], c["segmento"]])

# 3) ERP_PRODUTOS.csv — Latin-1, ';'
with open(f"{OUT}/ERP_PRODUTOS.csv", "w", encoding="latin-1", newline="") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(["COD_PROD","DESCRICAO","EAN","GRUPO","UNIDADE"])
    for p in produtos:
        w.writerow([p["cod_prod"], p["descricao"], p["ean"], p["categoria"], p["unidade"]])

# 4) ERP_TITULOS.TXT — arquivo POSICIONAL (largura fixa), Latin-1
#    Layout: FILIAL(2) NUMNOTA(6) PARCELA(2) COD_CLI(6) EMISSAO(8 ddmmaaaa)
#            VENCTO(8) VALOR(15, centavos, zero à esq.) STATUS(1: A/B)
tit = (erp_df.assign(v=erp_df["VLRTOTAL"].str.replace(",", ".").astype(float))
       .groupby(["FILIAL","NUMNOTA","CODCLI","EMISSAO"], as_index=False)["v"].sum())
with open(f"{OUT}/ERP_TITULOS.TXT", "w", encoding="latin-1", newline="\n") as f:
    f.write(f"{'BOASAFRA':<10}{'TITULOS A RECEBER':<30}{FIM.strftime('%d%m%Y')}\n")  # header
    for _, r in tit.iterrows():
        emiss = r["EMISSAO"].replace("/", "")
        dd, mm, aa = emiss[:2], emiss[2:4], emiss[4:]
        vc = date(int(aa), int(mm), int(dd)) + timedelta(days=30)
        cent = int(round(r["v"]*100))
        status = "B" if RNG.random() < 0.8 else "A"   # B=baixado(pago) A=aberto
        f.write(f"{r['FILIAL']:<2}{r['NUMNOTA']:<6}01{r['CODCLI']:<6}"
                f"{emiss:<8}{vc.strftime('%d%m%Y'):<8}{cent:015d}{status}\n")

# 5) forca_vendas.xlsx — Excel com abas Pedidos e Visitas (Title Case)
wb = Workbook()
ws = wb.active; ws.title = "Pedidos"
cols = ["Pedido","Data","CLIENTE_ID","CNPJ","Cliente","Representante","Produto",
        "Qtde","Preço Unit.","Total","Comissão %","UF"]
ws.append(cols)
for r in app_pedidos:
    ws.append([r[c] for c in cols])
ws.freeze_panes = "A2"
# Visitas (atividade dos reps, muitas sem venda -> custo)
wv = wb.create_sheet("Visitas")
wv.append(["Data","Representante","CLIENTE_ID","Cliente","Resultado","Observação"])
reps_cli = [c for c in clientes if "rep" in c["canais"]]
for _ in range(4200):
    c = reps_cli[RNG.integers(0, len(reps_cli))]
    dv = INI + timedelta(days=int(RNG.integers(0, (FIM-INI).days)))
    res = RNG.choice(["Pedido realizado","Sem pedido","Cliente ausente","Cobrança","Prospecção"],
                     p=[.45,.28,.12,.1,.05])
    wv.append([dv, c["rep"], c["cliente_id_app"], c["fantasia"], res, ""])
# cabeçalho em negrito
for sheet in (ws, wv):
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14463B")
wb.save(f"{OUT}/forca_vendas.xlsx")

# 6) ecommerce_pedidos.csv — UTF-8, ',', decimal '.', ISO date, snake_case
ecom_df = pd.DataFrame(ecom_linhas)
ecom_df.to_csv(f"{OUT}/ecommerce_pedidos.csv", index=False, encoding="utf-8")

# 7) planilhas_operacao.xlsx — as planilhas MANUAIS, bagunçadas
wb2 = Workbook()
# aba Custo_Produto (chave = descrição/código digitado, custo mensal, com furos)
w1 = wb2.active; w1.title = "Custo_Produto"
w1.append(["Planilha de custos - atualizar todo mês!", None, None, None])
w1.append(["Cod Produto","Descrição","Custo Compra (R$)","Últ. Atualização"])
meses = pd.date_range(INI, FIM, freq="MS")
for p in produtos:
    if RNG.random() < 0.05:            # 5% dos produtos nem estão na planilha (furo)
        continue
    fator = float(RNG.uniform(0.96, 1.06))
    custo = round(p["custo"]*fator, 2)
    cod = p["cod_prod"] if RNG.random() < 0.85 else ""     # às vezes só a descrição
    w1.append([cod, p["descricao"], money_br(custo),
               pd.Timestamp(RNG.choice(meses)).strftime("%m/%Y")])
# aba Frete_Cliente (chave = NOME DIGITADO À MÃO -> pesadelo de match)
w2 = wb2.create_sheet("Frete_Cliente")
w2.append(["Frete médio por cliente (levantado pelo Compras)"])
w2.append(["Cliente","Cidade","Frete por Entrega (R$)","Nº Entregas/Mês"])
for c in clientes:
    if RNG.random() < 0.12:            # nem todo cliente está mapeado
        continue
    frete = round(float(RNG.uniform(35, 180)) * (1.4 if c["uf"] not in ("MG","GO") else 1.0), 2)
    entregas = int(RNG.integers(1, 9))
    w2.append([baguncar_nome(c["razao"], c["fantasia"]), c["municipio"],
               money_br(frete), entregas])
# aba Metas_Comissao (por representante)
w3 = wb2.create_sheet("Metas_Comissao")
w3.append(["Metas e comissão - Comercial 2025/2026"])
w3.append(["Representante","Meta Mensal (R$)","Comissão Base %","Bônus Meta %"])
for rep in REPS:
    w3.append([rep, money_br(round(float(RNG.uniform(80000, 220000)), 2)),
               round(float(RNG.uniform(1.5, 3.0)), 2), round(float(RNG.uniform(0.3, 1.0)), 2)])
for w in (w1, w2, w3):
    w["A1"].font = Font(bold=True, italic=True, color="B8730F")
wb2.save(f"{OUT}/planilhas_operacao.xlsx")

# ---------------------------------------------------------------------------
print("Arquivos gerados em", OUT, ":")
print(f"  ERP_FATURAMENTO.csv .... {len(erp_df):>7,} linhas  (Latin-1, ';', DD/MM/AAAA)")
print(f"  ERP_CLIENTES.csv ....... {len(clientes):>7,} clientes")
print(f"  ERP_PRODUTOS.csv ....... {len(produtos):>7,} produtos")
print(f"  ERP_TITULOS.TXT ........ {len(tit):>7,} títulos  (posicional/largura fixa)")
print(f"  forca_vendas.xlsx ...... {len(app_pedidos):>7,} pedidos + 4.200 visitas (2 abas)")
print(f"  ecommerce_pedidos.csv .. {len(ecom_df):>7,} pedidos  (UTF-8, ISO, snake_case)")
print(f"  planilhas_operacao.xlsx  custos + fretes + metas (3 abas, nomes à mão)")
print(f"\nClientes por canal: rep={sum('rep' in c['canais'] for c in clientes)}, "
      f"ecom={sum('ecom' in c['canais'] for c in clientes)}")
